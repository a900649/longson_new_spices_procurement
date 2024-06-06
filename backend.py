from datetime import datetime
import streamlit as st
import pandas as pd
from io import BytesIO
import openpyxl as op
from openpyxl.styles import Font ,NamedStyle,Alignment
import warnings
import my_function as my
import yaml
from yaml.loader import SafeLoader
import streamlit_authenticator as stauth
import variables as v
from azure.storage.blob import BlobServiceClient,ContainerClient
import os
import math
import load_info

warnings.filterwarnings("ignore")
st.set_page_config(page_title=v.page_title, page_icon=load_info.icon_image)

with open("user.yaml") as file:
    config = yaml.load(file, Loader=SafeLoader)

authenticator = stauth.Authenticate(
    config['credentials'],
    config['cookie']['name'],
    config['cookie']['key'],
    config['cookie']['expiry_days'],
    config['preauthorized'],

)

def add_formula_col(results_df,check_df):
    if len(results_df) == 0:
        return results_df,[]
    else:
        results_df.reset_index(inplace=True, drop=False)

        different_list = []

        for i in range(0, len(results_df)):
            row_id = results_df.loc[i,"RowID"]
            for r in range(0, len(check_df)):
                check_column_name = check_df.loc[r, "Check Column Name"]
                equal_column = check_df.loc[r, "Equal Column"]
                formula = check_df.loc[r, "Formula"]
                for col in results_df.columns:
                    value = str(results_df.loc[i, col])
                    if value == "nan":
                        value = "0"
                    try:
                        # % 處理方式
                        if "%" in value:
                            value = str(float(value.replace("%","").strip()) / 100)
                    except:
                        pass
                    formula = formula.replace("[[" + col + "]]", value)
                try:
                    new_value = eval(formula)
                except Exception as error:
                    print(error)
                    new_value = 0
                results_df.loc[i, check_column_name] = new_value
                keyin_value = results_df.loc[i, equal_column]

                try: # 可能有問題，非數字
                    if math.isclose(float(keyin_value), new_value) == False:
                        different_list.append([row_id,check_column_name])
                except Exception as error:
                    print(error)
                    different_list.append([row_id, check_column_name])

        results_df.set_index('RowID', inplace=True)
        return results_df,different_list

def download_supplier_quotation_form():

    st.markdown("""
                <style>
                .title-font {
                    font-size:20px;
                    line-height: 0.7;
                    font-weight:bold;
                }
                </style>
                """, unsafe_allow_html=True)

    st.title("Download")
    select_program = st.selectbox(label="Program", options=v.program_name_list)

    if st.button('Confirm', key="4546"):
        # 讀取Info Data
        if v.data_source == "Excel":
            title_df, introduction_df, common_col_df, individual_col_df, function_df, verification_code_df, internal_mail_receipients_df, supplier_mail_setting_df, photo_df, check_df = my.load_info_by_excel()
        elif v.data_source == "Azure SQL":
            title_df, introduction_df, common_col_df, individual_col_df, function_df, verification_code_df, internal_mail_receipients_df, supplier_mail_setting_df, photo_df, check_df = my.load_info_by_sql(select_program)

        v.program_name = select_program
        v.results_file_path = "Procurement Results {}.xlsx".format(select_program)
        v.temp_path = "Temp" + "/" + select_program
        v.attachment_path = "Attachment" + "/" + select_program

        # 修改Info
        load_info.title_df = title_df
        load_info.introduction_df = introduction_df
        load_info.common_col_df = common_col_df
        load_info.individual_col_df = individual_col_df
        load_info.verification_code_df = verification_code_df
        load_info.check_df = check_df

    st.markdown('---')

    if st.session_state["authentication_status"]:
        st.markdown("<font class='title-font'>Selected Program : {}</font>".format(v.program_name),unsafe_allow_html=True)
        st.write("")
        title_len = len(load_info.title_df)
        introduction_len = 1

        c1, c2, c3= st.columns([1, 1, 1])

        with c1:
            st.write("Row Data")
            if st.button('Execute', key="1233"):
                product_list = list(load_info.introduction_df["Product"])

                #############################################################################################################
                # Azure 拉出來抓資料，就不用抓兩次
                if v.data_source == "Azure SQL":
                    results_df = my.load_results_table(v.program_name)
                    results_df, different_list = add_formula_col(results_df, load_info.check_df)
                #############################################################################################################

                wb = op.Workbook()
                for ind in range(0,len(product_list)):
                    product = product_list[ind]
                    sheetname = product.replace("/", " ").replace("\\", " ").replace("?", " ").replace("*", " ").replace("[", " ").replace("]", " ")  # Excel Sheetname 不接受
                    ws = wb.create_sheet(title=sheetname, index=ind)

                    #############################################################################################################
                    if v.data_source == "Excel":
                        try:
                            results_df = pd.read_excel(v.results_file_path, sheet_name=sheetname, index_col=0, engine="openpyxl")
                        except:
                            results_df = pd.DataFrame(columns=[])
                        results_df, different_list = add_formula_col(results_df, load_info.check_df)
                    #############################################################################################################
                    for i in range(0,title_len):
                        ws.cell(i+1, 1).value = load_info.title_df.loc[i,"Title"]

                    ws.cell(1 + title_len, 1).value = str(load_info.introduction_df.loc[ind, "Item"])

                    title_introduction_len = title_len + introduction_len + 1 # 多格一行

                    product_result_df = results_df[results_df["Product"] == product]

                    # 複製標題
                    for col in range(0, len(product_result_df.columns)):
                        ws.cell(1 + title_introduction_len, col + 1).value = product_result_df.columns[col]
                    # 複製資料
                    for ind in range(0, len(product_result_df)):
                        for col in range(0, len(product_result_df.columns)):
                            ws.cell(ind + 1 + 1 + title_introduction_len, col + 1).value = product_result_df.iloc[ind, col]

                    # 調整寬度
                    for i in range(0, len(list(ws.columns))):
                        col = list(ws.columns)[i]
                        column = col[0].column_letter  # Get the column name
                        if i == 0:
                            ws.column_dimensions[column].width = 80  # 先設定固定長度
                        else:
                            ws.column_dimensions[column].width = 15

                towrite = BytesIO()
                wb.save(towrite)
                towrite.seek(0)  # reset pointer
                st.download_button(label="Download", data=towrite.getvalue(), file_name="Row Data.xlsx", mime="application/vnd.ms-excel")



        ########################################

        with c2:
            st.write("Compared Form")
            if st.button('Execute', key="123"):
                product_list = list(load_info.introduction_df["Product"])

                #############################################################################################################
                # Azure 拉出來抓資料，就不用抓兩次
                if v.data_source == "Azure SQL":
                    results_df = my.load_results_table(v.program_name)
                    if load_info.use_verification_code:
                        results_df = my.keep_new_data(results_df, ["Product","Verification Code", "Verification Code Name"])
                    results_df, different_list = add_formula_col(results_df, load_info.check_df)
                #############################################################################################################

                sty1 = NamedStyle(name="sty1", font=Font(name='Calibra', size=11, color='000000', bold=True), alignment=Alignment(wrap_text=True, horizontal='left', vertical='center'))
                sty2 = NamedStyle(name="sty2", font=Font(name='Calibra', size=11, color='000000'), alignment=Alignment(wrap_text=True, horizontal='center', vertical='center'))
                sty3 = NamedStyle(name="sty3", font=Font(name='Calibra', size=11, color='000000'), alignment=Alignment(wrap_text=True, horizontal='left', vertical='center'))
                sty4 = NamedStyle(name="sty4", font=Font(name='Calibra', size=11, color='FF0000'), alignment=Alignment(wrap_text=True, horizontal='center', vertical='center'))

                wb = op.Workbook()
                for ind in range(0,len(product_list)):
                    product = product_list[ind]
                    sheetname = product.replace("/", " ").replace("\\", " ").replace("?", " ").replace("*", " ").replace("[", " ").replace("]", " ")  # Excel Sheetname 不接受
                    ws = wb.create_sheet(title=sheetname, index=ind)

                    #############################################################################################################
                    if v.data_source == "Excel":
                        try:
                            results_df = pd.read_excel(v.results_file_path, sheet_name=sheetname, index_col=0, engine="openpyxl")
                        except:
                            results_df = pd.DataFrame(columns=[])
                        if load_info.use_verification_code:
                            results_df = my.keep_new_data(results_df, ["Product", "Verification Code", "Verification Code Name"])
                        results_df, different_list = add_formula_col(results_df, load_info.check_df)
                    #############################################################################################################

                    campared_df = results_df[results_df["Product"] == product].T


                    for i in range(0, title_len):
                        ws.cell(i + 1, 1).value = load_info.title_df.loc[i, "Title"]
                        ws.cell(i + 1, 1).style = sty3

                    ws.cell(1 + title_len, 1).value = str(load_info.introduction_df.loc[ind, "Item"])
                    ws.cell(1 + title_len, 1).style = sty3

                    title_introduction_len = title_len + introduction_len + 1  # 多格一行

                    # 複製標題
                    for ind in range(0, len(results_df.columns)):
                        ws.cell(ind + 1 + title_introduction_len, 1).value = results_df.columns[ind]
                        ws.cell(ind + 1 + title_introduction_len, 1).style = sty1

                    datetime_type = type(datetime.today().date())

                    # 複製資料
                    for ind in range(0, len(campared_df)):
                        col_name = campared_df.index[ind]

                        for col in range(0, len(campared_df.columns)):
                            row_id = campared_df.columns[col]
                            values = campared_df.iloc[ind, col]

                            if col_name == "Update DateTime":
                                values = str(values)
                            if type(values) == datetime_type:
                                values = str(values)
                            ws.cell(ind + 1 + title_introduction_len, col + 1 + 1).value = values

                            if col_name in [i[1] for i in different_list if row_id == i[0]]:
                                # Check Data 不符合
                                ws.cell(ind + 1 + title_introduction_len, col + 1 + 1).style = sty4
                            else:
                                ws.cell(ind + 1 + title_introduction_len, col + 1 + 1).style = sty2

                    # 調整寬度
                    for i in range(0,len(list(ws.columns))):
                        col = list(ws.columns)[i]
                        column = col[0].column_letter  # Get the column name
                        if i == 0:
                            ws.column_dimensions[column].width = 80 # 先設定固定長度
                        else:
                            ws.column_dimensions[column].width = 20

                towrite = BytesIO()
                wb.save(towrite)
                towrite.seek(0)  # reset pointer
                st.download_button(label="Download", data=towrite.getvalue(), file_name="Compared.xlsx", mime="application/vnd.ms-excel")

        with c3:
            st.write("Download Attachment")

            if st.button('Execute', key="12"):

                # 若存在BLOB，先下載更新
                if v.data_source == "Azure SQL":
                    blob_service_client = BlobServiceClient.from_connection_string(v.blob_connection_string)
                    blob_container = ContainerClient.from_connection_string(conn_str=v.blob_connection_string, container_name=v.blob_container)
                    blob_list = blob_container.list_blobs("Attachment/{}".format(select_program))
                    for blob in blob_list:
                        print(blob.name)
                        # 檢查資料夾並建立資料夾
                        folder_list = blob.name.split("/")
                        for i in range(0,len(folder_list)-1):
                            folder_path = "/".join(folder_list[0:i+1])
                            if os.path.exists(folder_path) == False:
                                os.mkdir(folder_path)
                        # 檢查檔案並下載檔案
                        blob_client = blob_service_client.get_blob_client(container=v.blob_container, blob=blob.name)
                        if os.path.exists(blob.name) == False:
                            with open(file=blob.name, mode="wb") as file_blob:
                                download_stream = blob_client.download_blob()
                                file_blob.write(download_stream.readall())


                # 壓縮檔案
                my.attachment_file_zip(v.program_name)
                with open("Attachment.zip", "rb") as fp:
                    btn = st.download_button(
                        label="Download ZIP",
                        data=fp,
                        file_name="Attachment.zip",
                        mime="application/zip"
                    )
        ########################################

def login():

    if st.session_state["authentication_status"]:
        authenticator.logout()
    elif st.session_state["authentication_status"] is False:
        authenticator.login()
        st.error('Username/password is incorrect')
    elif st.session_state["authentication_status"] is None:
        authenticator.login()


page_names_to_funcs = {"Login":login,"Download": download_supplier_quotation_form}

functions_name = st.sidebar.selectbox("Option", page_names_to_funcs.keys())

page_names_to_funcs[functions_name]()
